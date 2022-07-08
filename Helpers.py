##########################
## ALL Helper Functions ##
##########################
import datetime
import os
from time import sleep
import pandas as pd

from PyQt5.QtWidgets import QMessageBox, QFileDialog
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Side, Border, Font, PatternFill
from openpyxl.utils import get_column_letter

current_date = datetime.date.today().strftime("%Y%m%d")

ran1wb = ""
ran2wb = ""


def gui_initialization(self):
    self.setWindowTitle("BB utilization tool")
    self.label_8.hide()

    self.textEdit.setEnabled(False)
    self.progressBar_ran1.hide()
    self.progressBar_ran2.hide()
    self.text_ran1_load.hide()
    self.text_ran2_load.hide()
    self.processing_files.hide()
    self.progressBar_process.hide()
    self.status_text.setText("Progress")
    self.status_text.hide()
    self.open_btn.setEnabled(False)


def messageBox(self, caption, message):
    QMessageBox.about(self, caption, message)


def load_files(path):
    print("Loading Files Started")
    try:
        wb = pd.read_excel(path, 'RSRAN131 - Node B Utilization')
        header_list = list(wb.columns.values)
        if 'Max BB SUs Util ratio' in header_list:
            print("File Loaded Successfully")
        else:
            print("Check input file")
        return wb
    except:
        messageBox(None, "Caution!", "File Not recognized")


def utilization_category(row):
    if row['Max Utilization'] > 90 or row['Max Utilization'] == 90:
        return 'Above 90%'
    if 0 <= row['Max Utilization'] < 60:
        return 'Below 60%'
    if 60 <= row['Max Utilization'] < 90:
        return 'Between 60%-90%'

    return 'Other'


def utilization_severity(row):
    if row['Repetition'] == 0:
        return 'No Severity'
    if row['Repetition'] == 1 or row['Repetition'] == 2:
        return 'Weak'
    if row['Repetition'] == 3 or row['Repetition'] == 4:
        return 'Moderate'
    if row['Repetition'] == 5 or row['Repetition'] == 6:
        return 'High'
    if row['Repetition'] == 7:
        return 'Very High'
    return 'Other'


def progressBar(self, x):
    self.completed = 0
    while self.completed < 100:
        self.completed += 1
        x.setValue(self.completed)
        sleep(0.03)


def prepare_sheets(self, ran1wb, ran2wb):
    print("Processing Started....")
    self.status_text.show()
    self.processing_files.show()
    self.processing_files.setText("Processing Files ....")
    self.status_text.show()
    self.status_text.setText("Progress")
    self.progressBar_process.show()
    self.progressBar_process.setValue(0)
    self.progressBar_process.setValue(6)
    print("Checking Headers....")
    self.processing_files.setText("Processing Files ....\nChecking Files Validity")
    ran1wb = ran1wb.dropna(subset=['PLMN name', 'Period start time'])
    ran2wb = ran2wb.dropna(subset=['PLMN name', 'Period start time'])
    print("Checking Headers Done!")
    self.processing_files.setText("Processing Files ....\nChecking Files Validity \nFiles Valid")
    self.progressBar_process.setValue(11)

    print("Cleaning ....")
    self.processing_files.setText("Processing Files ....\nChecking Files Validity \nFiles Valid \nCleaning Files")
    ran1wb = ran1wb[['Period start time', 'RNC name', 'WBTS name', 'WBTS ID',
                     'Max BB SUs Util ratio', 'NUMBER OF SUBUNITS IN BASEBAND']]
    ran1wb['WBTS ID'] = ran1wb['WBTS ID'].astype(int)
    ran2wb = ran2wb[['Period start time', 'RNC name', 'WBTS name', 'WBTS ID',
                     'Max BB SUs Util ratio', 'NUMBER OF SUBUNITS IN BASEBAND']]
    ran2wb['WBTS ID'] = ran2wb['WBTS ID'].astype(int)
    self.progressBar_process.setValue(23)
    print("Cleaning extra Rows/Columns Done!")
    self.processing_files.setText("Processing Files ....\nChecking Files Validity \nFiles Valid \n"
                                  "Cleaning Files \nCleaning extra Rows/Columns Done!")
    print("Appending files....")
    self.processing_files.setText("Processing Files ....\nChecking Files Validity \nFiles Valid \n"
                                  "Cleaning Files \nCleaning extra Rows/Columns Done! \n"
                                  "Appending files")
    list_of_files = [ran1wb, ran2wb]
    append_ran = pd.concat(list_of_files)
    append_ran.insert(0, "Region", append_ran['WBTS name'].str[-2:])
    append_ran.insert(3, "WBTS Code", append_ran['WBTS name'].str[-6:])
    append_ran.insert(4, "RNC", append_ran['RNC name'].str[-2:])
    append_ran.insert(0, "Key", append_ran['RNC'].astype(str) + "_" + append_ran['WBTS ID'].astype(str))
    append_ran = append_ran.drop(append_ran.columns[[3]], axis=1)
    append_ran['Period start time'] = pd.to_datetime(append_ran['Period start time']).dt.date
    self.progressBar_process.setValue(59)
    print("Appending files Done....")
    self.processing_files.setText("Processing Files ....\nChecking Files Validity \nFiles Valid \n"
                                  "Cleaning Files \nCleaning extra Rows/Columns Done! \n"
                                  "Appending files \nAppending Files Done!")
    print("Creating Pivot table....")
    self.processing_files.setText("Processing Files ....\nChecking Files Validity \nFiles Valid \n"
                                  "Cleaning Files \nCleaning extra Rows/Columns Done! \n"
                                  "Appending files \nAppending Files Done! \nCreating Pivot table")
    table = pd.pivot_table(append_ran, values='Max BB SUs Util ratio', index=['Key', 'WBTS name', 'Region', 'RNC'],
                           columns=['Period start time']).reset_index()
    self.progressBar_process.setValue(66)
    print("Generating Maximum utilization this week")
    self.processing_files.setText("Processing Files ....\nChecking Files Validity \nFiles Valid \n"
                                  "Cleaning Files \nCleaning extra Rows/Columns Done! \n"
                                  "Appending files \nAppending Files Done! \nCreating Pivot table \n"
                                  "Generating Maximum utilization this week")
    table["Max Utilization"] = table.max(numeric_only=True, axis=1)
    self.progressBar_process.setValue(69)
    print("Calculating Repetitions....")
    self.processing_files.setText("Processing Files ....\nChecking Files Validity \nFiles Valid \n"
                                  "Cleaning Files \nCleaning extra Rows/Columns Done! \n"
                                  "Appending files \nAppending Files Done! \nCreating Pivot table \n"
                                  "Generating Maximum utilization this week \nCalculating Repetitions")
    table["Repetition"] = table.iloc[:, -8:-1].gt(89.999999).sum(axis=1)
    self.progressBar_process.setValue(70)
    print("Creating Category/Severity ...")
    self.processing_files.setText("Processing Files ....\nChecking Files Validity \nFiles Valid \n"
                                  "Cleaning Files \nCleaning extra Rows/Columns Done! \n"
                                  "Appending files \nAppending Files Done! \nCreating Pivot table \n"
                                  "Generating Maximum utilization this week \nCalculating Repetitions \n"
                                  "Creating Category/Severity")
    table["Category"] = table.apply(lambda row: utilization_category(row), axis=1)
    table["Severity"] = table.apply(lambda row: utilization_severity(row), axis=1)
    self.progressBar_process.setValue(71)
    print("Creating Category/Severity Done.")
    mapper = lambda x: x.strftime("%d-%m-%Y") if isinstance(x, datetime.datetime) else x
    table.columns = table.columns.map(mapper)
    self.progressBar_process.setValue(72)
    print("Pivot table done.")
    self.processing_files.setText("Processing Files ....\nChecking Files Validity \nFiles Valid \n"
                                  "Cleaning Files \nCleaning extra Rows/Columns Done! \n"
                                  "Appending files \nAppending Files Done! \nCreating Pivot table \n"
                                  "Generating Maximum utilization this week \nCalculating Repetitions \n"
                                  "Creating Category/Severity \nPivot Table done \nCreating Summary")
    print("Creating Summary....")
    self.progressBar_process.setValue(78)
    nodeB_count = table.shape[0]
    above_90_count = table['Category'].value_counts()['Above 90%']
    between_60_90_count = table['Category'].value_counts()['Between 60%-90%']
    below_60_count = table['Category'].value_counts()['Below 60%']
    noSeverity_count = table['Severity'].value_counts()['No Severity']
    weak_moderate_count = table['Severity'].value_counts()['Weak'] + \
                          table['Severity'].value_counts()['Moderate']
    high_veryhigh_count = table['Severity'].value_counts()['High'] + \
                          table['Severity'].value_counts()['Very High']
    above_90_per = float("{0:.2f}".format(above_90_count / (table.shape[0]) * 100))
    between_60_90_per = float("{0:.2f}".format((between_60_90_count / (table.shape[0])) * 100))
    below_60_per = float("{0:.2f}".format((below_60_count / (table.shape[0])) * 100))
    noSeverity_per = float("{0:.2f}".format((noSeverity_count / (table.shape[0])) * 100))
    weak_moderate_per = float("{0:.2f}".format((weak_moderate_count / (table.shape[0])) * 100))
    high_veryhigh_per = float("{0:.2f}".format((high_veryhigh_count / (table.shape[0])) * 100))

    summary_dict = {
        'Key': ['Count of Node B', 'Above 90', 'Between 60% and 90%', 'Below 60%', "High Severity", "No Severity",
                "Weak/Moderate"],
        'Value': [nodeB_count, above_90_count, between_60_90_count, below_60_count, noSeverity_count,
                  weak_moderate_count, high_veryhigh_count],
        'Percentage': ['100%', above_90_per, between_60_90_per, below_60_per, noSeverity_per, weak_moderate_per,
                       high_veryhigh_per]
    }
    summary_df = pd.DataFrame(summary_dict)

    self.processing_files.setText("Processing Files ....\nChecking Files Validity \nFiles Valid \n"
                                  "Cleaning Files \nCleaning extra Rows/Columns Done! \n"
                                  "Appending files \nAppending Files Done! \nCreating Pivot table \n"
                                  "Generating Maximum utilization this week \nCalculating Repetitions \n"
                                  "Creating Category/Severity \nPivot Table done")

    cell_format(self, saveReport(self, append_ran, table, summary_df))

    self.processing_files.setText("Processing Files ....\nChecking Files Validity \nFiles Valid \n"
                                  "Cleaning Files \nCleaning extra Rows/Columns Done! \n"
                                  "Appending files \nAppending Files Done! \nCreating Pivot table \n"
                                  "Generating Maximum utilization this week \nCalculating Repetitions \n"
                                  "Creating Category/Severity \nPivot Table done \nFormatting cells \n"
                                  "Alignment cells/Borders \nAlignment cells/Borders Done \n"
                                  "Auto fit columns \nAuto fit columns Done\nConditional formatting \n"
                                  "Conditional formatting Done \nCells Formatting Done \nCreating Summary \n"
                                  "BB Utilization Sheet is Ready!")

    self.label_8.show()
    self.label_8.setText("#NBs:  " + str(nodeB_count))
    self.below_60.setText(str(below_60_count))
    self.between_60_90.setText(str(between_60_90_count))
    self.above_90.setText(str(above_90_count))
    self.no_severity.setText(str(noSeverity_count))
    self.weak_moderate.setText(str(weak_moderate_count))
    self.high_veryhigh.setText(str(high_veryhigh_count))
    self.progressBar_process.setValue(100)
    self.open_btn.setEnabled(True)
    print("BB utilization sheet is now ready!")


def cell_format(self, file_path_1):
    self.progressBar_process.setValue(80)
    print("Formatting cells ..")
    print("Alignment cells / Borders ..")
    self.processing_files.setText("Processing Files ....\nChecking Files Validity \nFiles Valid \n"
                                  "Cleaning Files \nCleaning extra Rows/Columns Done! \n"
                                  "Appending files \nAppending Files Done! \nCreating Pivot table \n"
                                  "Generating Maximum utilization this week \nCalculating Repetitions \n"
                                  "Creating Category/Severity \nPivot Table done \nFormatting cells \n"
                                  "Alignment cells/Borders")
    wb = load_workbook(file_path_1)
    ws = wb[wb.sheetnames[1]]
    analysis_sheet = wb['Analysis']
    for sheet in wb.worksheets:
        for row_cells in sheet.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment(wrapText=None, horizontal='center', vertical='center')
    for row_cells in analysis_sheet.iter_rows():
        for cell in row_cells:
            thin = Side(border_style="thin", color="00000000")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    print("Alignment cells / Borders Done")
    self.processing_files.setText("Processing Files ....\nChecking Files Validity \nFiles Valid \n"
                                  "Cleaning Files \nCleaning extra Rows/Columns Done! \n"
                                  "Appending files \nAppending Files Done! \nCreating Pivot table \n"
                                  "Generating Maximum utilization this week \nCalculating Repetitions \n"
                                  "Creating Category/Severity \nPivot Table done \nFormatting cells \n"
                                  "Alignment cells/Borders \nAlignment cells/Borders Done ")
    self.progressBar_process.setValue(93)

    print("Auto fit columns ..")
    self.processing_files.setText("Processing Files ....\nChecking Files Validity \nFiles Valid \n"
                                  "Cleaning Files \nCleaning extra Rows/Columns Done! \n"
                                  "Appending files \nAppending Files Done! \nCreating Pivot table \n"
                                  "Generating Maximum utilization this week \nCalculating Repetitions \n"
                                  "Creating Category/Severity \nPivot Table done \nFormatting cells \n"
                                  "Alignment cells/Borders \nAlignment cells/Borders Done \n"
                                  "Auto fit columns ")
    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            new_column_length = max(len(str(cell.value)) for cell in column_cells)
            new_column_letter = (get_column_letter(column_cells[0].column))
            if new_column_length > 0:
                sheet.column_dimensions[new_column_letter].width = new_column_length * 1.02
    print("Auto fit columns Done")
    self.processing_files.setText("Processing Files ....\nChecking Files Validity \nFiles Valid \n"
                                  "Cleaning Files \nCleaning extra Rows/Columns Done! \n"
                                  "Appending files \nAppending Files Done! \nCreating Pivot table \n"
                                  "Generating Maximum utilization this week \nCalculating Repetitions \n"
                                  "Creating Category/Severity \nPivot Table done \nFormatting cells \n"
                                  "Alignment cells/Borders \nAlignment cells/Borders Done \n"
                                  "Auto fit columns \nAuto fit columns Done")
    self.progressBar_process.setValue(96)

    print("Conditional formatting .. ")
    self.processing_files.setText("Processing Files ....\nChecking Files Validity \nFiles Valid \n"
                                  "Cleaning Files \nCleaning extra Rows/Columns Done! \n"
                                  "Appending files \nAppending Files Done! \nCreating Pivot table \n"
                                  "Generating Maximum utilization this week \nCalculating Repetitions \n"
                                  "Creating Category/Severity \nPivot Table done \nFormatting cells \n"
                                  "Alignment cells/Borders \nAlignment cells/Borders Done \n"
                                  "Auto fit columns \nAuto fit columns Done\nConditional formatting ")
    for rows in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=12):
        for cell in rows:
            if cell.value is None:
                cell.value = 0
            elif int(cell.value) >= 90:
                cell.value = float("{0:.2f}".format(cell.value))
                cell.font = Font(color='FF0000')
                cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            elif 83 < int(cell.value) < 90:
                cell.value = float("{0:.2f}".format(cell.value))
                cell.font = Font(color='FF8000')
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            elif int(cell.value) <= 83:
                cell.value = float("{0:.2f}".format(cell.value))
                cell.font = Font(color='009900')
                cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    for rows in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in rows:
            cell.font = Font(color='000000')
            cell.fill = PatternFill(start_color="FF8000", end_color="FF8000", fill_type="solid")
    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
        for cell in rows:
            cell.font = Font(color='000000')
            cell.fill = PatternFill(start_color="FF8000", end_color="FF8000", fill_type="solid")
    print("Conditional formatting Done ")
    self.processing_files.setText("Processing Files ....\nChecking Files Validity \nFiles Valid \n"
                                  "Cleaning Files \nCleaning extra Rows/Columns Done! \n"
                                  "Appending files \nAppending Files Done! \nCreating Pivot table \n"
                                  "Generating Maximum utilization this week \nCalculating Repetitions \n"
                                  "Creating Category/Severity \nPivot Table done \nFormatting cells \n"
                                  "Alignment cells/Borders \nAlignment cells/Borders Done \n"
                                  "Auto fit columns \nAuto fit columns Done\nConditional formatting \n"
                                  "Conditional formatting Done \nCells Formatting Done ")

    wb.save(file_path_1)

    print("Formatting cells Done")


def saveReport(self, df1, df2, df3):
    response = QFileDialog.getSaveFileName(caption='Save Report', directory=f'AppendedBB_{current_date}',
                                           filter="Excel (*.xlsx *.xls *.csv)")

    if response[0] == "":
        QMessageBox.about(self, 'Caution', "Please specify save location!")
    else:
        writer = pd.ExcelWriter(response[0], engine='openpyxl')
        df1.to_excel(writer, sheet_name='Raw Data', index=False)
        df2.to_excel(writer, sheet_name='Analysis', index=False)
        df3.to_excel(writer, sheet_name='Summary', index=False)
        writer.save()

    return response[0]
